#!/usr/bin/env python3
"""
Import Enriched Excel to Google Sheets

This script imports the SC_Ethics_Districts_Enriched.xlsx file
to the SC Ethics Monitor Google Sheet.

Supports two modes:
1. API Mode: Uses Google Sheets API with service account credentials
2. Manual Mode: Provides instructions for manual import

Prerequisites for API Mode:
- Google service account credentials JSON file
- Set GOOGLE_SHEETS_CREDENTIALS environment variable or place credentials.json in project root

Usage:
    python import_to_sheets.py                    # Auto-detect mode
    python import_to_sheets.py --manual           # Force manual instructions
    python import_to_sheets.py --credentials path/to/creds.json
"""

import argparse
import os
import sys
from pathlib import Path

# Project paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / "data"
EXCEL_FILE = DATA_DIR / "SC_Ethics_Districts_Enriched.xlsx"

# Google Sheet details
SPREADSHEET_ID = "17j_KFZFUw-ESBQlKlIccUMpGCFq_XdeL6WYph7zkxQo"
SPREADSHEET_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/edit"


def check_excel_file():
    """Check if the Excel file exists and return info about it."""
    if not EXCEL_FILE.exists():
        return None, "Excel file not found. Run generate_enriched_excel.py first."

    file_size = EXCEL_FILE.stat().st_size
    modified = EXCEL_FILE.stat().st_mtime
    from datetime import datetime
    modified_str = datetime.fromtimestamp(modified).strftime("%Y-%m-%d %H:%M:%S")

    return {
        "path": str(EXCEL_FILE),
        "size_bytes": file_size,
        "size_kb": round(file_size / 1024, 1),
        "modified": modified_str,
    }, None


def find_credentials():
    """Find Google Sheets credentials file."""
    # Check environment variable
    env_creds = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    if env_creds and Path(env_creds).exists():
        return Path(env_creds)

    # Check common locations
    locations = [
        PROJECT_ROOT / "credentials.json",
        PROJECT_ROOT / "google-credentials.json",
        PROJECT_ROOT / "service-account.json",
        Path.home() / ".config" / "gspread" / "credentials.json",
        Path.home() / ".config" / "gspread" / "service_account.json",
    ]

    for loc in locations:
        if loc.exists():
            return loc

    return None


def import_via_api(credentials_path: Path):
    """Import Excel data to Google Sheets via API."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        from openpyxl import load_workbook
    except ImportError as e:
        print(f"Missing dependency: {e}")
        print("Install with: pip install gspread google-auth openpyxl")
        return False

    print(f"\nLoading Excel file: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, data_only=False)
    print(f"  Worksheets found: {wb.sheetnames}")

    print(f"\nConnecting to Google Sheets...")
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
    ]
    creds = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
    client = gspread.authorize(creds)

    try:
        spreadsheet = client.open_by_key(SPREADSHEET_ID)
        print(f"  Connected to: {spreadsheet.title}")
    except Exception as e:
        print(f"\nError: Could not access spreadsheet: {e}")
        print("\nMake sure the service account has edit access to the spreadsheet.")
        print(f"Share the sheet with the service account email found in: {credentials_path}")
        return False

    # Import each worksheet
    for sheet_name in ['Dashboard', 'Districts', 'Lists']:
        if sheet_name not in wb.sheetnames:
            print(f"  Skipping {sheet_name} (not in Excel file)")
            continue

        print(f"\nImporting {sheet_name} tab...")
        ws_excel = wb[sheet_name]

        # Get data from Excel
        data = []
        for row in ws_excel.iter_rows(values_only=True):
            # Convert None to empty string and handle formulas
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                elif isinstance(cell, (int, float)):
                    row_data.append(cell)
                else:
                    row_data.append(str(cell))
            data.append(row_data)

        if not data:
            print(f"  No data in {sheet_name}")
            continue

        # Get or create the worksheet in Google Sheets
        try:
            ws_gsheet = spreadsheet.worksheet(sheet_name)
            print(f"  Found existing worksheet: {sheet_name}")
            ws_gsheet.clear()
        except gspread.WorksheetNotFound:
            print(f"  Creating new worksheet: {sheet_name}")
            ws_gsheet = spreadsheet.add_worksheet(
                title=sheet_name,
                rows=len(data) + 10,
                cols=max(len(row) for row in data) + 5
            )

        # Write data
        print(f"  Writing {len(data)} rows...")
        ws_gsheet.update(data, value_input_option='USER_ENTERED')
        print(f"  Done!")

    # Reorder sheets (Dashboard first)
    try:
        dashboard = spreadsheet.worksheet("Dashboard")
        spreadsheet.reorder_worksheets([dashboard] + [
            w for w in spreadsheet.worksheets() if w.title != "Dashboard"
        ])
    except Exception:
        pass  # Not critical

    print("\n" + "=" * 60)
    print("SUCCESS! Data imported to Google Sheets")
    print("=" * 60)
    print(f"\nView your sheet: {SPREADSHEET_URL}")

    return True


def show_manual_instructions(excel_info: dict):
    """Show instructions for manual import."""
    print("\n" + "=" * 60)
    print("MANUAL IMPORT INSTRUCTIONS")
    print("=" * 60)

    print("\nStep 1: Locate the Excel File")
    print("-" * 40)
    print(f"  Path: {excel_info['path']}")
    print(f"  Size: {excel_info['size_kb']} KB")
    print(f"  Modified: {excel_info['modified']}")

    print("\nStep 2: Open Google Sheets")
    print("-" * 40)
    print(f"  URL: {SPREADSHEET_URL}")
    print("  (Click this link or copy/paste into browser)")

    print("\nStep 3: Import the Excel File")
    print("-" * 40)
    print("  1. Click 'File' in the menu bar")
    print("  2. Select 'Import'")
    print("  3. Click the 'Upload' tab")
    print("  4. Drag the Excel file into the upload area")
    print("     OR click 'Select a file from your device'")
    print("  5. Choose 'Replace spreadsheet' for the import location")
    print("  6. Click 'Import data'")

    print("\nStep 4: Verify Import")
    print("-" * 40)
    print("  After import, verify:")
    print("  - Dashboard tab shows 18 KPIs (blue tab)")
    print("  - Districts tab has 170 rows with 22 columns (green tab)")
    print("  - Conditional formatting shows party colors (blue D/red R)")
    print("  - Lists tab exists (may be hidden)")

    print("\nExpected Data Summary:")
    print("-" * 40)
    print("  - 170 districts (124 House + 46 Senate)")
    print("  - Party breakdown: 121 R, 48 D, 1 Open")
    print("  - Regions: Upstate, Midlands, Lowcountry, Pee Dee")
    print("  - 31 competitive districts (< 10% margin)")

    print("\n" + "=" * 60)
    print("TIP: To open Finder at the file location, run:")
    print(f"  open \"{excel_info['path'].rsplit('/', 1)[0]}\"")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(description="Import Excel to Google Sheets")
    parser.add_argument("--manual", action="store_true", help="Show manual import instructions")
    parser.add_argument("--credentials", type=str, help="Path to Google credentials JSON")
    args = parser.parse_args()

    print("=" * 60)
    print("SC Ethics Monitor - Import to Google Sheets")
    print("=" * 60)

    # Check Excel file
    excel_info, error = check_excel_file()
    if error:
        print(f"\nError: {error}")
        print("\nRun the Excel generator first:")
        print("  cd sc-ethics-monitor")
        print("  python scripts/generate_enriched_excel.py")
        sys.exit(1)

    print(f"\nExcel file found:")
    print(f"  Path: {excel_info['path']}")
    print(f"  Size: {excel_info['size_kb']} KB")
    print(f"  Modified: {excel_info['modified']}")

    # Check for manual mode
    if args.manual:
        show_manual_instructions(excel_info)
        return

    # Find credentials
    credentials_path = None
    if args.credentials:
        credentials_path = Path(args.credentials)
        if not credentials_path.exists():
            print(f"\nError: Credentials file not found: {credentials_path}")
            sys.exit(1)
    else:
        credentials_path = find_credentials()

    if credentials_path:
        print(f"\nCredentials found: {credentials_path}")
        print("Attempting API import...")

        success = import_via_api(credentials_path)
        if not success:
            print("\nAPI import failed. Showing manual instructions instead.")
            show_manual_instructions(excel_info)
    else:
        print("\nNo Google credentials found.")
        print("To enable API import, either:")
        print("  1. Set GOOGLE_SHEETS_CREDENTIALS environment variable")
        print("  2. Place credentials.json in the project root")
        print("  3. Use --credentials flag to specify path")
        show_manual_instructions(excel_info)


if __name__ == "__main__":
    main()
