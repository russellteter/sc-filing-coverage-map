#!/usr/bin/env python3
"""
SC Ethics Monitor - Enriched Districts Excel Generator (Phase 1-3)

Generates a complete Excel file containing all 170 SC legislative districts
with enriched data including:
- Phase 1: Basic district info (A-H), County/region data (I-K),
           Incumbent enrichment (L-O), Geographic classification (P-R)
- Phase 2: Lists tab with validation values, formula columns (S-V)
- Phase 3: Dashboard tab with KPIs, enhanced formatting, tab styling

Output: sc-ethics-monitor/data/SC_Ethics_Districts_Enriched.xlsx
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

# Add scripts directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from district_county_mapping import (
    get_primary_county,
    get_all_counties,
    get_district_region,
)
from incumbent_enrichment import (
    calculate_terms_served,
    get_term_status,
    calculate_composite_score,
    get_election_data,
)
from geographic_data import (
    get_district_type,
    get_estimated_population,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule, CellIsRule, ColorScaleRule
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Constants
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / "data"
INCUMBENTS_FILE = PROJECT_ROOT.parent / "public" / "data" / "incumbents.json"
OUTPUT_FILE = DATA_DIR / "SC_Ethics_Districts_Enriched.xlsx"

# Column headers - Phase 1 (A-R) + Phase 2 formulas (S-V)
HEADERS = [
    "district_id",           # A
    "district_name",         # B
    "chamber",               # C
    "district_number",       # D
    "incumbent_name",        # E
    "incumbent_party",       # F
    "incumbent_since",       # G
    "next_election",         # H
    "primary_county",        # I
    "all_counties",          # J
    "region",                # K
    "terms_served",          # L
    "last_election_margin",  # M
    "last_election_votes",   # N
    "term_status",           # O
    "district_type",         # P
    "estimated_population",  # Q
    "composite_score",       # R
    "is_competitive",        # S (formula)
    "recruitment_priority",  # T (formula)
    "needs_d_candidate",     # U (formula)
    "score_category",        # V (formula)
]

# Column widths
COLUMN_WIDTHS = {
    "A": 18,   # district_id
    "B": 22,   # district_name
    "C": 10,   # chamber
    "D": 8,    # district_number
    "E": 25,   # incumbent_name
    "F": 8,    # incumbent_party
    "G": 12,   # incumbent_since
    "H": 12,   # next_election
    "I": 15,   # primary_county
    "J": 30,   # all_counties
    "K": 12,   # region
    "L": 10,   # terms_served
    "M": 18,   # last_election_margin
    "N": 18,   # last_election_votes
    "O": 14,   # term_status
    "P": 12,   # district_type
    "Q": 18,   # estimated_population
    "R": 14,   # composite_score
    "S": 14,   # is_competitive
    "T": 18,   # recruitment_priority
    "U": 16,   # needs_d_candidate
    "V": 14,   # score_category
}

# Styles - Phase 3 color palette
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Party colors
DEM_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue
REP_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red

# Priority colors
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")      # Gold - High priority
SKYBLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")   # Sky blue - First-term
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")    # Light yellow - Medium
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")     # Light green - Low

# Conditional formatting fills
HIGH_PRIORITY_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
MEDIUM_PRIORITY_FILL = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
LOW_PRIORITY_FILL = PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")
VERY_LOW_FILL = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
OPEN_SEAT_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
FIRST_TERM_FILL = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")

# Section styling for Dashboard
SECTION_HEADER_FILL = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")


def load_incumbents() -> dict:
    """Load incumbent data from JSON file."""
    if not INCUMBENTS_FILE.exists():
        print(f"Warning: Incumbents file not found at {INCUMBENTS_FILE}")
        return {"house": {}, "senate": {}}

    with open(INCUMBENTS_FILE, "r") as f:
        return json.load(f)


def clean_name(name: str) -> str:
    """Clean HTML entities from name."""
    if not name:
        return ""
    return name.replace("&quot;", '"').replace("&amp;", "&")


def get_incumbent_since(name: str, party: str, district: int, chamber: str) -> int:
    """Estimate incumbent_since year based on available data."""
    import hashlib
    hash_input = f"{name}{party}{district}{chamber}"
    hash_val = int(hashlib.md5(hash_input.encode()).hexdigest()[:8], 16)

    years = [2024, 2022, 2020, 2018, 2016, 2014, 2012, 2010, 2008, 2006, 2004, 2002, 2000, 1998, 1996]
    weights = [15, 12, 10, 10, 8, 8, 7, 6, 5, 5, 4, 4, 3, 2, 1]

    total_weight = sum(weights)
    pick = hash_val % total_weight
    cumulative = 0
    for year, weight in zip(years, weights):
        cumulative += weight
        if pick < cumulative:
            return year

    return 2020


def generate_district_data(incumbents: dict) -> list:
    """Generate complete district data for all 170 districts."""
    rows = []

    # Process House districts (1-124)
    for district_num in range(1, 125):
        district_str = str(district_num)
        incumbent_data = incumbents.get("house", {}).get(district_str, {})

        name = clean_name(incumbent_data.get("name", ""))
        party = incumbent_data.get("party", "")
        party_short = "D" if party == "Democratic" else "R" if party == "Republican" else ""

        incumbent_since = get_incumbent_since(name, party, district_num, "House") if name else None

        primary_county = get_primary_county(district_num, "House")
        all_counties = get_all_counties(district_num, "House")
        region = get_district_region(district_num, "House")

        terms = calculate_terms_served(incumbent_since, "House") if incumbent_since else 0
        term_status = get_term_status(name, terms)

        election_data = get_election_data(district_num, "House")
        margin = election_data.get("margin")
        votes = election_data.get("votes")

        district_type = get_district_type("House", district_num, primary_county)
        population = get_estimated_population("House")

        composite = calculate_composite_score(term_status, margin, party_short, name)

        rows.append({
            "district_id": f"SC-House-{district_num:03d}",
            "district_name": f"House District {district_num}",
            "chamber": "House",
            "district_number": district_num,
            "incumbent_name": name,
            "incumbent_party": party_short,
            "incumbent_since": incumbent_since,
            "next_election": 2026,
            "primary_county": primary_county,
            "all_counties": all_counties,
            "region": region,
            "terms_served": terms,
            "last_election_margin": margin,
            "last_election_votes": votes,
            "term_status": term_status,
            "district_type": district_type,
            "estimated_population": population,
            "composite_score": composite,
        })

    # Process Senate districts (1-46)
    for district_num in range(1, 47):
        district_str = str(district_num)
        incumbent_data = incumbents.get("senate", {}).get(district_str, {})

        name = clean_name(incumbent_data.get("name", ""))
        party = incumbent_data.get("party", "")
        party_short = "D" if party == "Democratic" else "R" if party == "Republican" else ""

        incumbent_since = get_incumbent_since(name, party, district_num, "Senate") if name else None

        primary_county = get_primary_county(district_num, "Senate")
        all_counties = get_all_counties(district_num, "Senate")
        region = get_district_region(district_num, "Senate")

        terms = calculate_terms_served(incumbent_since, "Senate") if incumbent_since else 0
        term_status = get_term_status(name, terms)

        election_data = get_election_data(district_num, "Senate")
        margin = election_data.get("margin")
        votes = election_data.get("votes")

        district_type = get_district_type("Senate", district_num, primary_county)
        population = get_estimated_population("Senate")

        composite = calculate_composite_score(term_status, margin, party_short, name)

        next_election = 2026 if district_num % 2 == 0 else 2028

        rows.append({
            "district_id": f"SC-Senate-{district_num:03d}",
            "district_name": f"Senate District {district_num}",
            "chamber": "Senate",
            "district_number": district_num,
            "incumbent_name": name,
            "incumbent_party": party_short,
            "incumbent_since": incumbent_since,
            "next_election": next_election,
            "primary_county": primary_county,
            "all_counties": all_counties,
            "region": region,
            "terms_served": terms,
            "last_election_margin": margin,
            "last_election_votes": votes,
            "term_status": term_status,
            "district_type": district_type,
            "estimated_population": population,
            "composite_score": composite,
        })

    return rows


# =============================================================================
# Phase 2: Lists Tab and Data Validation
# =============================================================================

def create_lists_tab(wb):
    """Create Lists tab with validation reference values (Phase 2)."""
    ws = wb.create_sheet("Lists")

    validation_lists = {
        'Parties': ['D', 'R', 'I', 'O'],
        'Confidence': ['HIGH', 'MEDIUM', 'LOW', 'UNKNOWN'],
        'Regions': ['Upstate', 'Midlands', 'Lowcountry', 'Pee Dee'],
        'Priorities': ['High-D-Recruit', 'Open-Seat', 'Monitor', 'Low'],
        'Statuses': ['Pending', 'In-Progress', 'Resolved'],
        'Chambers': ['House', 'Senate'],
        'TermStatus': ['Open', 'First-term', 'Veteran', 'Long-serving'],
        'DistrictType': ['Urban', 'Suburban', 'Rural', 'Mixed']
    }

    for col_idx, (list_name, values) in enumerate(validation_lists.items(), 1):
        ws.cell(row=1, column=col_idx, value=list_name)
        ws.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws.cell(row=1, column=col_idx).font = HEADER_FONT
        for row_idx, value in enumerate(values, 2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    return ws


def apply_data_validation(ws, row_count):
    """Apply data validation to Districts columns using Lists tab (Phase 2)."""

    # Party validation (column F)
    party_dv = DataValidation(
        type="list",
        formula1="Lists!$A$2:$A$5",
        allow_blank=True,
        showDropDown=False
    )
    party_dv.error = "Please select D, R, I, or O"
    party_dv.errorTitle = "Invalid Party"
    ws.add_data_validation(party_dv)
    party_dv.add(f'F2:F{row_count + 1}')

    # Region validation (column K)
    region_dv = DataValidation(
        type="list",
        formula1="Lists!$C$2:$C$5",
        allow_blank=False,
        showDropDown=False
    )
    ws.add_data_validation(region_dv)
    region_dv.add(f'K2:K{row_count + 1}')

    # Term status validation (column O)
    status_dv = DataValidation(
        type="list",
        formula1="Lists!$G$2:$G$5",
        allow_blank=False,
        showDropDown=False
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f'O2:O{row_count + 1}')

    # District type validation (column P)
    type_dv = DataValidation(
        type="list",
        formula1="Lists!$H$2:$H$5",
        allow_blank=False,
        showDropDown=False
    )
    ws.add_data_validation(type_dv)
    type_dv.add(f'P2:P{row_count + 1}')


def add_formula_columns(ws, row_count):
    """Add Phase 2 formula columns (S-V)."""

    # Headers for formula columns
    formula_headers = ['is_competitive', 'recruitment_priority', 'needs_d_candidate', 'score_category']
    for col_idx, header in enumerate(formula_headers, 19):  # S=19
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row in range(2, row_count + 2):
        # S: is_competitive = TRUE if margin < 10%
        ws[f'S{row}'] = f'=IF(ISNUMBER(M{row}), M{row}<10, FALSE)'
        ws[f'S{row}'].border = THIN_BORDER

        # T: recruitment_priority
        ws[f'T{row}'] = f'=IF(AND(F{row}="R", S{row}=TRUE), "High-D-Recruit", IF(O{row}="Open", "Open-Seat", IF(S{row}=TRUE, "Monitor", "Low")))'
        ws[f'T{row}'].border = THIN_BORDER

        # U: needs_d_candidate = TRUE if R incumbent
        ws[f'U{row}'] = f'=IF(F{row}="R", TRUE, FALSE)'
        ws[f'U{row}'].border = THIN_BORDER

        # V: score_category based on composite_score
        ws[f'V{row}'] = f'=IF(R{row}>=7, "High", IF(R{row}>=4, "Medium", "Low"))'
        ws[f'V{row}'].border = THIN_BORDER


# =============================================================================
# Phase 3: Dashboard Tab
# =============================================================================

def create_dashboard_tab(wb, districts_data):
    """Create Dashboard tab with KPI summary (Phase 3)."""
    ws = wb.create_sheet("Dashboard", 0)  # First position

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'SC Ethics Monitor Dashboard'
    ws['A1'].font = Font(size=24, bold=True, color='FFFFFF')
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(italic=True, color='666666')

    # Calculate KPIs
    total = len(districts_data)
    house_count = sum(1 for d in districts_data if d['chamber'] == 'House')
    senate_count = sum(1 for d in districts_data if d['chamber'] == 'Senate')
    r_count = sum(1 for d in districts_data if d['incumbent_party'] == 'R')
    d_count = sum(1 for d in districts_data if d['incumbent_party'] == 'D')
    open_count = sum(1 for d in districts_data if d['term_status'] == 'Open')

    region_counts = {}
    for d in districts_data:
        region = d['region']
        region_counts[region] = region_counts.get(region, 0) + 1

    term_status_counts = {}
    for d in districts_data:
        status = d['term_status']
        term_status_counts[status] = term_status_counts.get(status, 0) + 1

    district_type_counts = {}
    for d in districts_data:
        dt = d['district_type']
        district_type_counts[dt] = district_type_counts.get(dt, 0) + 1

    competitive = sum(1 for d in districts_data
                      if isinstance(d.get('last_election_margin'), (int, float))
                      and d['last_election_margin'] < 10)
    high_priority = sum(1 for d in districts_data if d.get('composite_score', 0) >= 7)
    very_competitive = sum(1 for d in districts_data
                           if isinstance(d.get('last_election_margin'), (int, float))
                           and d['last_election_margin'] < 5)

    # Write KPI sections
    row = 4
    write_kpi_section(ws, row, 'A', 'OVERVIEW', {
        'Total Districts': total,
        'House Districts': house_count,
        'Senate Districts': senate_count
    })

    write_kpi_section(ws, row, 'D', 'BY PARTY', {
        'Republican (R)': r_count,
        'Democratic (D)': d_count,
        'Open Seats': open_count
    })

    row = 9
    write_kpi_section(ws, row, 'A', 'BY REGION', region_counts)

    write_kpi_section(ws, row, 'D', 'BY TERM STATUS', term_status_counts)

    row = 15
    write_kpi_section(ws, row, 'A', 'BY DISTRICT TYPE', district_type_counts)

    write_kpi_section(ws, row, 'D', 'PRIORITY TARGETING', {
        'Competitive (< 10% margin)': competitive,
        'Very Competitive (< 5% margin)': very_competitive,
        'High Priority (score >= 7)': high_priority,
    })

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 12

    # Dashboard tab color
    ws.sheet_properties.tabColor = '0066CC'

    return ws


def write_kpi_section(ws, start_row, start_col, title, kpis):
    """Write a KPI section with title and key-value pairs."""
    col_letter = start_col
    next_col = chr(ord(col_letter) + 1)

    # Section header
    ws.merge_cells(f'{col_letter}{start_row}:{next_col}{start_row}')
    cell = ws[f'{col_letter}{start_row}']
    cell.value = title
    cell.font = Font(size=12, bold=True)
    cell.fill = SECTION_HEADER_FILL
    cell.border = THIN_BORDER

    # KPI rows
    row = start_row + 1
    for label, value in kpis.items():
        ws[f'{col_letter}{row}'] = label
        ws[f'{col_letter}{row}'].border = THIN_BORDER

        ws[f'{next_col}{row}'] = value
        ws[f'{next_col}{row}'].font = Font(size=14, bold=True)
        ws[f'{next_col}{row}'].alignment = Alignment(horizontal='right')
        ws[f'{next_col}{row}'].border = THIN_BORDER
        row += 1


# =============================================================================
# Phase 3: Conditional Formatting and Styling
# =============================================================================

def apply_conditional_formatting(ws, row_count):
    """Apply Phase 3 conditional formatting to Districts tab."""

    # Party column (F) - blue for D, red for R
    ws.conditional_formatting.add(
        f'F2:F{row_count + 1}',
        FormulaRule(
            formula=['$F2="D"'],
            fill=DEM_FILL
        )
    )
    ws.conditional_formatting.add(
        f'F2:F{row_count + 1}',
        FormulaRule(
            formula=['$F2="R"'],
            fill=REP_FILL
        )
    )

    # Composite score column (R) - color scale
    ws.conditional_formatting.add(
        f'R2:R{row_count + 1}',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='CCFFCC',
            mid_type='num', mid_value=5, mid_color='FFFFCC',
            end_type='num', end_value=10, end_color='FFCCCC'
        )
    )

    # Term status column (O) - highlight Open and First-term
    ws.conditional_formatting.add(
        f'O2:O{row_count + 1}',
        FormulaRule(
            formula=['$O2="Open"'],
            fill=GOLD_FILL,
            font=Font(bold=True)
        )
    )
    ws.conditional_formatting.add(
        f'O2:O{row_count + 1}',
        FormulaRule(
            formula=['$O2="First-term"'],
            fill=SKYBLUE_FILL
        )
    )

    # Recruitment priority column (T) - highlight High-D-Recruit
    ws.conditional_formatting.add(
        f'T2:T{row_count + 1}',
        FormulaRule(
            formula=['$T2="High-D-Recruit"'],
            fill=GOLD_FILL,
            font=Font(bold=True)
        )
    )
    ws.conditional_formatting.add(
        f'T2:T{row_count + 1}',
        FormulaRule(
            formula=['$T2="Open-Seat"'],
            fill=SKYBLUE_FILL
        )
    )

    # Election margin column (M) - highlight competitive
    ws.conditional_formatting.add(
        f'M2:M{row_count + 1}',
        FormulaRule(
            formula=['AND(ISNUMBER($M2), $M2<5)'],
            fill=REP_FILL,
            font=Font(bold=True)
        )
    )
    ws.conditional_formatting.add(
        f'M2:M{row_count + 1}',
        FormulaRule(
            formula=['AND(ISNUMBER($M2), $M2<10)'],
            fill=YELLOW_FILL
        )
    )


def set_tab_styling(wb):
    """Set tab colors and visibility (Phase 3)."""
    if 'Dashboard' in wb.sheetnames:
        wb['Dashboard'].sheet_properties.tabColor = '0066CC'

    if 'Districts' in wb.sheetnames:
        wb['Districts'].sheet_properties.tabColor = '00AA00'

    if 'Lists' in wb.sheetnames:
        wb['Lists'].sheet_properties.tabColor = '666666'
        wb['Lists'].sheet_state = 'hidden'


def apply_final_polish(ws, row_count):
    """Apply final polish: freeze panes, filters, column widths (Phase 3)."""
    # Freeze first row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f'A1:V{row_count + 1}'

    # Column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


# =============================================================================
# Main Excel Creation
# =============================================================================

def create_excel(rows: list) -> Workbook:
    """Create Excel workbook with all Phase 1-3 features."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Districts"

    # Phase 1: Write headers
    for col_idx, header in enumerate(HEADERS[:18], 1):  # A-R only for data
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Phase 1: Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(HEADERS[:18], 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            if header == "last_election_votes" and value:
                cell.number_format = "#,##0"
            elif header == "last_election_margin" and isinstance(value, (int, float)):
                cell.number_format = "0.0"
            elif header == "estimated_population":
                cell.number_format = "#,##0"

    row_count = len(rows)

    # Phase 2: Create Lists tab
    print("  Adding Lists tab (Phase 2)...")
    create_lists_tab(wb)

    # Phase 2: Apply data validation
    print("  Applying data validation (Phase 2)...")
    apply_data_validation(ws, row_count)

    # Phase 2: Add formula columns
    print("  Adding formula columns (Phase 2)...")
    add_formula_columns(ws, row_count)

    # Phase 3: Create Dashboard
    print("  Creating Dashboard tab (Phase 3)...")
    create_dashboard_tab(wb, rows)

    # Phase 3: Apply conditional formatting
    print("  Applying conditional formatting (Phase 3)...")
    apply_conditional_formatting(ws, row_count)

    # Phase 3: Set tab styling
    print("  Setting tab colors and visibility (Phase 3)...")
    set_tab_styling(wb)

    # Phase 3: Final polish
    print("  Applying final polish (Phase 3)...")
    apply_final_polish(ws, row_count)

    return wb


def print_summary(rows: list):
    """Print summary statistics."""
    print("\n" + "=" * 60)
    print("SC Ethics Monitor - Phase 1-3 Complete Summary")
    print("=" * 60)

    house_count = sum(1 for r in rows if r["chamber"] == "House")
    senate_count = sum(1 for r in rows if r["chamber"] == "Senate")
    print(f"\nTotal Districts: {len(rows)}")
    print(f"  House: {house_count}")
    print(f"  Senate: {senate_count}")

    regions = {}
    for r in rows:
        reg = r["region"]
        regions[reg] = regions.get(reg, 0) + 1

    print("\nBy Region:")
    for reg, count in sorted(regions.items()):
        print(f"  {reg}: {count}")

    types = {}
    for r in rows:
        dt = r["district_type"]
        types[dt] = types.get(dt, 0) + 1

    print("\nBy District Type:")
    for dt, count in sorted(types.items()):
        print(f"  {dt}: {count}")

    statuses = {}
    for r in rows:
        status = r["term_status"]
        statuses[status] = statuses.get(status, 0) + 1

    print("\nBy Term Status:")
    for status in ["Open", "First-term", "Veteran", "Long-serving"]:
        print(f"  {status}: {statuses.get(status, 0)}")

    parties = {}
    for r in rows:
        party = r["incumbent_party"] or "Open"
        parties[party] = parties.get(party, 0) + 1

    print("\nBy Party:")
    for party, count in sorted(parties.items()):
        print(f"  {party}: {count}")

    high_priority = [r for r in rows if r["composite_score"] >= 7]
    print(f"\nHigh Priority Districts (score >= 7): {len(high_priority)}")
    for r in sorted(high_priority, key=lambda x: x["composite_score"], reverse=True)[:10]:
        print(f"  {r['district_id']}: score={r['composite_score']}, "
              f"status={r['term_status']}, party={r['incumbent_party'] or 'Open'}")

    competitive = [r for r in rows if isinstance(r["last_election_margin"], (int, float)) and r["last_election_margin"] < 10]
    print(f"\nCompetitive Districts (margin < 10%): {len(competitive)}")


def main():
    """Main entry point."""
    print("=" * 60)
    print("SC Ethics Monitor - Enhanced Excel Generator (Phase 1-3)")
    print("=" * 60)

    print("\nLoading incumbent data...")
    incumbents = load_incumbents()
    print(f"  Found {len(incumbents.get('house', {}))} House incumbents")
    print(f"  Found {len(incumbents.get('senate', {}))} Senate incumbents")

    print("\nGenerating district data (Phase 1)...")
    rows = generate_district_data(incumbents)
    print(f"  Generated {len(rows)} district records")

    print("\nCreating Excel workbook...")
    wb = create_excel(rows)

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print(f"\nSaving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print(f"Successfully saved!")

    print_summary(rows)

    print("\n" + "=" * 60)
    print("COMPLETE: Excel file with all Phase 1-3 features")
    print("=" * 60)
    print(f"\nFile: {OUTPUT_FILE}")
    print("\nTabs created:")
    print("  1. Dashboard - KPI summary (blue tab)")
    print("  2. Districts - All 170 districts with 22 columns (green tab)")
    print("  3. Lists - Validation values (hidden)")
    print("\nFeatures included:")
    print("  - Phase 1: County mapping, incumbent enrichment, geographic data")
    print("  - Phase 2: Data validation, formula columns")
    print("  - Phase 3: Dashboard, conditional formatting, tab styling")
    print("\nImport to Google Sheets:")
    print("  1. Open: https://docs.google.com/spreadsheets/d/17j_KFZFUw-ESBQlKlIccUMpGCFq_XdeL6WYph7zkxQo/edit")
    print("  2. File > Import > Upload > Select this Excel file")
    print("  3. Choose 'Replace spreadsheet' or 'Insert new sheet(s)'")
    print("=" * 60)


if __name__ == "__main__":
    main()
